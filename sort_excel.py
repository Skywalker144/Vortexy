#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取xlsx文件夹中的每个Excel文件，通过噪声-温度拟合曲线计算41dBA对应的温度，
按温度从小到大排序，并为每个文件保存对应的排序后的Excel文件。
"""

import os
import numpy as np
import pandas as pd
from scipy import interpolate
from openpyxl import Workbook


def read_fan_data(file_path):
    """
    读取Excel文件中的风扇数据
    返回: {风扇型号: [(噪声, 温度, 转速), ...], ...}
    """
    df = pd.read_excel(file_path, header=None)
    
    fan_data = {}
    
    # 第一行是风扇型号，每3列为一组
    header_row = df.iloc[0]
    
    col = 0
    while col < len(header_row):
        fan_name = header_row[col]
        
        # 跳过空列
        if pd.isna(fan_name) or str(fan_name).strip() == '':
            col += 1
            continue
        
        fan_name = str(fan_name).strip()
        
        # 读取该风扇的数据（噪声、温度、转速三列）
        data_points = []
        for row_idx in range(1, len(df)):
            try:
                noise = df.iloc[row_idx, col]
                temp = df.iloc[row_idx, col + 1]
                rpm = df.iloc[row_idx, col + 2]
                
                # 检查数据是否有效
                if pd.notna(noise) and pd.notna(temp) and pd.notna(rpm):
                    noise = float(noise)
                    temp = float(temp)
                    rpm = float(rpm)
                    data_points.append((noise, temp, rpm))
            except (ValueError, IndexError):
                continue
        
        if data_points:
            fan_data[fan_name] = data_points
        
        col += 3  # 移动到下一个风扇
    
    return fan_data


def calculate_temp_at_41dba(data_points):
    """
    通过噪声-温度拟合曲线计算41dBA对应的温度
    使用线性插值
    """
    if len(data_points) < 2:
        return None
    
    # 按噪声排序
    sorted_points = sorted(data_points, key=lambda x: x[0])
    
    noises = [p[0] for p in sorted_points]
    temps = [p[1] for p in sorted_points]
    
    # 检查41dBA是否在数据范围内
    min_noise = min(noises)
    max_noise = max(noises)
    
    target_noise = 41.0
    
    # 如果41dBA在范围内，使用插值
    if min_noise <= target_noise <= max_noise:
        # 使用线性插值
        f = interpolate.interp1d(noises, temps, kind='linear')
        return float(f(target_noise))
    elif target_noise < min_noise:
        # 如果41dBA小于最小噪声，使用外推（基于前两个点）
        if len(noises) >= 2:
            # 使用线性外推
            slope = (temps[1] - temps[0]) / (noises[1] - noises[0])
            return temps[0] + slope * (target_noise - noises[0])
        return None
    else:
        # 如果41dBA大于最大噪声，使用外推（基于最后两个点）
        if len(noises) >= 2:
            slope = (temps[-1] - temps[-2]) / (noises[-1] - noises[-2])
            return temps[-1] + slope * (target_noise - noises[-1])
        return None


def process_single_file(input_path, output_path):
    """
    处理单个Excel文件，按41dBA温度排序后保存
    """
    print(f"处理文件: {input_path}")
    
    fan_data = read_fan_data(input_path)
    
    # 计算每个风扇的41dBA温度
    fans_with_temp = []
    for fan_name, data_points in fan_data.items():
        temp_41dba = calculate_temp_at_41dba(data_points)
        
        if temp_41dba is not None:
            fans_with_temp.append({
                'name': fan_name,
                'temp_41dba': temp_41dba,
                'data_points': data_points
            })
            print(f"  {fan_name}: {temp_41dba:.2f}°C @ 41dBA")
        else:
            print(f"  {fan_name}: 无法计算41dBA温度，跳过")
    
    if not fans_with_temp:
        print(f"  没有有效数据，跳过文件")
        return
    
    # 按41dBA温度从小到大排序
    fans_with_temp.sort(key=lambda x: x['temp_41dba'])
    
    # 创建输出Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入数据，保持原始格式（每个风扇3列：噪声、温度、转速）
    col = 1
    for fan in fans_with_temp:
        # 写入风扇型号（第一行）
        ws.cell(row=1, column=col, value=fan['name'])
        
        # 写入数据点
        for row_idx, (noise, temp, rpm) in enumerate(fan['data_points'], start=2):
            ws.cell(row=row_idx, column=col, value=noise)
            ws.cell(row=row_idx, column=col + 1, value=temp)
            ws.cell(row=row_idx, column=col + 2, value=rpm)
        
        col += 3  # 移动到下一个风扇的位置
    
    # 保存文件
    wb.save(output_path)
    print(f"  排序结果已保存到: {output_path}")
    print(f"  排序顺序: {' < '.join([f['name'] for f in fans_with_temp])}")
    print()


def main():
    xlsx_folder = 'xlsx_origin'
    output_folder = 'xlsx_sorted'
    
    # 创建输出文件夹
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    # 遍历xlsx文件夹中的所有Excel文件
    processed_count = 0
    for filename in os.listdir(xlsx_folder):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            input_path = os.path.join(xlsx_folder, filename)
            output_path = os.path.join(output_folder, filename)
            
            try:
                process_single_file(input_path, output_path)
                processed_count += 1
            except Exception as e:
                print(f"处理文件 {filename} 出错: {e}")
                print()
    
    print(f"完成！共处理 {processed_count} 个文件")
    print(f"排序后的文件保存在: {output_folder}/")


if __name__ == '__main__':
    main()
